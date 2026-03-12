"""
Excel/CSV parser for mill stock lists.

Supported layouts:
  - CAH two-column: species+product in col A, qty in col G,
                    species+product in col I, qty in col N
  - Standard table: header row with labeled columns
  - Fallback: col A as products with nearest numeric column as qty

All rows pass through the scoring pipeline before database insertion.
"""
import re
import logging
import openpyxl
import pandas as pd
from .base import ParseResult, ParsedRow
from .cleaning import is_noise_text, detect_mill_name
from .row_scoring import is_candidate_row
from lumber_normalizer import (
    normalize_thickness, normalize_grade, normalize_species,
    build_product_string
)

log = logging.getLogger("excel_parser")


# ═══════════════════════════════════════════════════════════════════════════════
# ENTRY POINT
# ═══════════════════════════════════════════════════════════════════════════════

def parse_excel(file_path: str) -> ParseResult:
    """Parse an Excel or CSV file into structured product rows."""
    result = ParseResult(parsing_method="excel")

    try:
        if file_path.lower().endswith('.csv'):
            return _parse_csv(file_path, result)

        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active

        layout = _detect_layout(ws)
        result.warnings.append(f"Detected layout: {layout}")
        log.info(f"Excel layout: {layout}")

        if layout == "cah_two_column":
            rows, mill_info = _parse_cah_two_column(ws)
        else:
            rows, mill_info = _parse_generic_table(ws)

        result.mill_name = mill_info.get("mill_name")
        result.mill_contact = mill_info.get("contact")
        result.mill_email = mill_info.get("email")
        result.mill_phone = mill_info.get("phone")
        result.mill_location = mill_info.get("location")

        # Store metadata for pipeline mill-name detection
        result._metadata = {
            'mill_name_candidates': [(0, mill_info['mill_name'])] if mill_info.get('mill_name') else [],
            'stock_date': None,
            'emails': [mill_info['email']] if mill_info.get('email') else [],
            'phones': [mill_info['phone']] if mill_info.get('phone') else [],
            'locations': [mill_info['location']] if mill_info.get('location') else [],
        }

        result.rows = [r for r in rows if r is not None]

        if result.rows:
            result.success = True
            result.confidence = 0.8
        else:
            result.errors.append("No product rows found in Excel file.")

    except Exception as e:
        log.error(f"Excel parsing error: {e}", exc_info=True)
        result.success = False
        result.errors.append(f"Excel parsing error: {str(e)}")

    return result


# ═══════════════════════════════════════════════════════════════════════════════
# LAYOUT DETECTION
# ═══════════════════════════════════════════════════════════════════════════════

def _detect_layout(ws) -> str:
    """Detect worksheet layout type."""
    used_cols = set()
    for row in ws.iter_rows(min_row=1, max_row=min(20, ws.max_row)):
        for cell in row:
            if cell.value is not None:
                used_cols.add(cell.column_letter)

    # CAH pattern: A + G + I + N, col G has numbers
    if 'A' in used_cols and 'G' in used_cols and 'I' in used_cols and 'N' in used_cols:
        g_vals = [ws.cell(row=r, column=7).value for r in range(9, min(20, ws.max_row))]
        g_nums = sum(1 for v in g_vals if isinstance(v, (int, float)))
        if g_nums >= 3:
            return "cah_two_column"

    return "standard_table"


# ═══════════════════════════════════════════════════════════════════════════════
# CAH TWO-COLUMN PARSER
# ═══════════════════════════════════════════════════════════════════════════════

def _parse_cah_two_column(ws) -> tuple:
    """
    Parse CAH layout:
      Col A = left product/species, Col G = left qty
      Col I = right product/species, Col N = right qty
      Col H = mill info in header rows
    """
    mill_info = {}
    rows = []

    # Mill info from col H header rows
    for r in range(1, 9):
        val = ws.cell(row=r, column=8).value
        if val and str(val).strip():
            v = str(val).strip()
            if not mill_info.get("mill_name") and re.search(
                    r'(hardwood|lumber|sawmill|wood|forest)', v, re.I):
                mill_info["mill_name"] = v
            elif not mill_info.get("mill_name") and r <= 2:
                mill_info["mill_name"] = v
            if re.search(r'@', v):
                mill_info["email"] = v
            if re.search(r'\d{3}[-.\s]\d{3}', v):
                mill_info["phone"] = v

    species_left = None
    species_right = None

    for r in range(9, ws.max_row + 1):
        val_a = _cell_str(ws.cell(row=r, column=1))
        val_g = ws.cell(row=r, column=7).value
        val_i = _cell_str(ws.cell(row=r, column=9))
        val_n = ws.cell(row=r, column=14).value

        if not val_a and not val_i and val_g is None and val_n is None:
            continue

        # Left column
        if val_a:
            if _is_species_header(val_a):
                species_left = normalize_species(val_a)
            elif not is_noise_text(val_a):
                row = _build_cah_row(val_a, val_g, species_left, r)
                if row:
                    rows.append(row)

        # Right column
        if val_i:
            if _is_species_header(val_i):
                species_right = normalize_species(val_i)
            elif not is_noise_text(val_i):
                row = _build_cah_row(val_i, val_n, species_right, r)
                if row:
                    rows.append(row)

    return rows, mill_info


def _build_cah_row(product_str: str, qty_val, species: str,
                   row_num: int) -> ParsedRow:
    """Build a ParsedRow from a CAH product string and quantity cell."""
    product_str = product_str.strip()
    if not product_str or len(product_str) < 3:
        return None

    # Skip obvious non-product rows
    if is_noise_text(product_str):
        return None

    row = ParsedRow(species=species, raw_text=product_str, source_row=row_num)

    # Qty
    if qty_val is not None and isinstance(qty_val, (int, float)):
        row.quantity_numeric = float(qty_val)
        row.quantity = f"{int(qty_val):,} BF"
    elif qty_val is not None:
        row.quantity = str(qty_val).strip()

    # Extract thickness
    m = re.match(r'^(\d+/\d+)\s+(.*)', product_str)
    if m:
        row.thickness = normalize_thickness(m.group(1))
        remainder = m.group(2).strip()
        _parse_cah_description(remainder, row)
    else:
        # No thickness — could still be valid if it has grade info
        _parse_cah_description(product_str, row)

    row.product = product_str
    row.product_normalized = build_product_string(row.to_dict())

    return row


def _parse_cah_description(text: str, row: ParsedRow):
    """Extract grade, color, surface, width from a CAH product description."""
    if not text:
        return

    original = text

    # Width constraints (8"&W, 12"&W, 10"-11")
    width_m = re.search(r'(\d+["\']?\s*&?\s*[Ww]|[\d.]+"\s*-\s*[\d."]+)', text)
    if width_m:
        row.width = width_m.group().strip()
        text = text[:width_m.start()].strip()

    # Surface
    surf_m = re.search(r'\b(15/16|13/16|S2S|S4S|RGH|Rough)\b', text, re.I)
    if surf_m:
        row.surface = surf_m.group()
        text = text[:surf_m.start()].strip() + " " + text[surf_m.end():].strip()
        text = text.strip()

    # Color
    color_m = re.search(r'\b(White|Brown|Red|Natural|Sap|Heart|Calico|Wht)\b', text, re.I)
    if color_m:
        row.color = color_m.group().title().replace('Wht', 'White')

    # Treatment
    treat_m = re.search(r'\b(Wormy|Ambrosia|Stained|Steamed|Rustic|PG)\b', text, re.I)
    if treat_m:
        row.treatment = treat_m.group()

    # Grade = remainder after stripping known components
    grade_text = re.sub(r'\b(15/16|13/16|S2S|S4S|RGH|Rough)\b', '', text, flags=re.I)
    grade_text = re.sub(r'\b(White|Brown|Red|Natural|Sap|Heart|Calico|Wht)\b', '', grade_text, flags=re.I)
    grade_text = re.sub(r'\d+["\']?\s*&?\s*[Ww]', '', grade_text)
    grade_text = re.sub(r'\s+', ' ', grade_text).strip()
    if grade_text:
        row.grade = normalize_grade(grade_text)

    row.description = original


# ═══════════════════════════════════════════════════════════════════════════════
# GENERIC TABLE PARSER
# ═══════════════════════════════════════════════════════════════════════════════

def _parse_generic_table(ws) -> tuple:
    """Parse a standard Excel table with header row."""
    mill_info = {}
    rows = []

    # Find header row
    header_row_idx = None
    col_map = {}

    for r in range(1, min(15, ws.max_row + 1)):
        row_vals = {c.column: str(c.value).strip().lower() for c in ws[r] if c.value}
        keywords_found = 0
        temp_map = {}
        for col_idx, val in row_vals.items():
            if any(k in val for k in ('species', 'product', 'description', 'grade', 'thick')):
                keywords_found += 1
                temp_map[val] = col_idx
            if any(k in val for k in ('qty', 'quantity', 'bf', 'bft', 'board feet')):
                temp_map['qty'] = col_idx
            if 'price' in val:
                temp_map['price'] = col_idx
            if 'length' in val or 'board length' in val:
                temp_map['length'] = col_idx
            if 'mill' in val:
                temp_map['mill'] = col_idx
        if keywords_found >= 2:
            header_row_idx = r
            col_map = temp_map
            break

    # Extract mill info from pre-header rows
    header_scan_limit = header_row_idx or min(12, ws.max_row)
    for r in range(1, header_scan_limit + 1):
        for cell in ws[r]:
            val = _cell_str(cell)
            if not val:
                continue
            if not mill_info.get('mill_name') and re.search(
                    r'(hardwood|lumber|sawmill|wood|forest|valley|timber)', val, re.I):
                mill_info['mill_name'] = val
            if not mill_info.get('mill_name') and r == 1 and len(val) > 4:
                mill_info['mill_name'] = val
            if re.search(r'@', val):
                mill_info['email'] = val
            if re.search(r'\d{3}[-.\s]\d{4}', val):
                mill_info['phone'] = val
            if re.search(r'[A-Z]{2}\s+\d{5}', val):
                mill_info['location'] = val

    # Parse with header
    if header_row_idx and col_map:
        current_species = None
        for r in range(header_row_idx + 1, ws.max_row + 1):
            row_data = {col: ws.cell(row=r, column=col).value
                        for col in range(1, ws.max_column + 1)}
            row = _parse_table_data_row(row_data, col_map, current_species, r)
            if row:
                if row.species and _is_species_header(row.species):
                    current_species = normalize_species(row.species)
                else:
                    rows.append(row)
        return rows, mill_info

    # Fallback: col A as products
    current_species = None
    for r in range(1, ws.max_row + 1):
        val_a = _cell_str(ws.cell(row=r, column=1))
        if not val_a:
            continue

        qty_val = None
        for c in range(2, min(ws.max_column + 1, 10)):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, (int, float)) and v > 0:
                qty_val = v
                break

        if _is_species_header(val_a):
            current_species = normalize_species(val_a)
            continue

        if is_noise_text(val_a):
            continue

        row = _build_cah_row(val_a, qty_val, current_species, r)
        if row:
            rows.append(row)

    return rows, mill_info


def _parse_table_data_row(row_data: dict, col_map: dict,
                           species: str, row_num: int) -> ParsedRow:
    """Parse a single data row from a table-style Excel sheet."""
    row = ParsedRow(species=species, source_row=row_num)

    def get(key, *fallback_keys):
        for k in [key] + list(fallback_keys):
            col = col_map.get(k)
            if col and row_data.get(col) is not None:
                val = str(row_data[col]).strip()
                if val and not is_noise_text(val):
                    return val
        return None

    def get_num(key):
        col = col_map.get(key)
        if col:
            v = row_data.get(col)
            if isinstance(v, (int, float)):
                return float(v)
        return None

    sp = get('species')
    if sp:
        row.species = normalize_species(sp)

    desc = get('description', 'product')
    if desc:
        row.description = desc
        row.product = desc
        m = re.match(r'^(\d+/\d+)\s+(.*)', desc)
        if m:
            row.thickness = normalize_thickness(m.group(1))
            row.grade = normalize_grade(m.group(2))
        else:
            row.grade = normalize_grade(desc)

    grade = get('grade')
    if grade:
        row.grade = normalize_grade(grade)

    thick = get('thick', 'thickness')
    if thick:
        row.thickness = normalize_thickness(thick)

    qty = get_num('qty') or get_num('quantity') or get_num('bf') or get_num('bft')
    if qty is not None:
        row.quantity_numeric = qty
        row.quantity = f"{int(qty):,} BF"
    else:
        qty_str = get('qty', 'quantity')
        if qty_str:
            row.quantity = qty_str

    price = get('price')
    if price:
        row.price = price

    length = get('length')
    if length:
        row.length = length

    if not row.product:
        parts = [v for v in [row.thickness, row.grade] if v]
        row.product = ' '.join(parts)

    if not row.product:
        return None

    row.product_normalized = build_product_string(row.to_dict())
    row.raw_text = str({k: row_data.get(v) for k, v in col_map.items()})
    return row


# ═══════════════════════════════════════════════════════════════════════════════
# CSV PARSER
# ═══════════════════════════════════════════════════════════════════════════════

def _parse_csv(file_path: str, result: ParseResult) -> ParseResult:
    """Parse a CSV file."""
    try:
        df = pd.read_csv(file_path)
        rows = []
        df.columns = [str(c).strip() for c in df.columns]

        for idx, r in df.iterrows():
            row = ParsedRow(source_row=idx)

            for col in df.columns:
                col_l = col.lower()
                val = r[col] if pd.notna(r[col]) else None
                if val is None:
                    continue
                val_str = str(val)

                if is_noise_text(val_str):
                    continue

                if 'species' in col_l or 'wood' in col_l:
                    row.species = normalize_species(val_str)
                elif 'thick' in col_l:
                    row.thickness = normalize_thickness(val_str)
                elif 'grade' in col_l:
                    row.grade = normalize_grade(val_str)
                elif any(k in col_l for k in ('qty', 'quantity', 'bf')):
                    try:
                        row.quantity_numeric = float(val_str.replace(',', ''))
                        row.quantity = f"{int(row.quantity_numeric):,} BF"
                    except ValueError:
                        row.quantity = val_str
                elif 'desc' in col_l or 'product' in col_l:
                    row.description = val_str
                    row.product = val_str
                elif 'price' in col_l:
                    row.price = val_str
                elif 'length' in col_l:
                    row.length = val_str

            if not row.product:
                parts = [v for v in [row.thickness, row.grade, row.description] if v]
                row.product = ' '.join(parts)

            if row.product:
                row.product_normalized = build_product_string(row.to_dict())
                rows.append(row)

        result.rows = rows
        result.success = bool(rows)
        result.confidence = 0.7
    except Exception as e:
        result.errors.append(f"CSV parsing error: {str(e)}")

    return result


# ═══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def _cell_str(cell) -> str:
    if cell.value is None:
        return ""
    return str(cell.value).strip()


def _is_species_header(text: str) -> bool:
    """Check if text is a lumber species header."""
    if not text:
        return False
    text = text.strip().rstrip(':').strip('*- ').lower()

    species_words = [
        'ash', 'aspen', 'basswood', 'beech', 'birch', 'cedar', 'cherry',
        'coffeenut', 'cottonwood', 'cypress', 'elm', 'hackberry', 'hemlock',
        'hickory', 'locust', 'maple', 'oak', 'pine', 'poplar',
        'sassafras', 'sycamore', 'walnut', 'alder', 'gum',
    ]
    for sw in species_words:
        if sw in text:
            if not re.search(r'\d/\d', text):
                return True
    return False
